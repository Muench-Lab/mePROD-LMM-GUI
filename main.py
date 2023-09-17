__author__ = "Süleyman Bozkurt"
__version__ = "v2.1"
__maintainer__ = "Süleyman Bozkurt"
__email__ = "sbozkurt.mbg@gmail.com"
__date__ = '18.01.2022'
__update__ = '17.09.2023'

import os
from threading import Thread
from tkinter import *
from tkinter.font import Font
from tkinter.scrolledtext import ScrolledText
from tkinter import filedialog, messagebox
from functions import *
from tkinter import END,Tk,Frame,Label,Button,StringVar
from openpyxl import Workbook
import openpyxl.styles as sty
from datetime import datetime
import re
import random

class MyWindow():
    def __init__(self, parent):
        self.filename_condition = ''
        self.frame = Frame(parent, width=840, height=560)
        self.font = Font(family="Times New Roman", size=16)
        self.request_timeout = 30
        global root
        root = parent

        # normalisation variables: Total intensity normalisation, Median intensity normalisation, TMM
        Label(self.frame, font=Font(family="Times New Roman", size=12), text="Normalization Method").place(x = 600, y = 5)

        self.fontRadio = Font(family="Times New Roman", size=13)
        self.normVar = StringVar()
        self.totalRadio = Radiobutton(root, font=self.fontRadio,  text="Total intensity", value="total", variable=self.normVar)
        self.totalRadio.select()
        self.totalRadio.place(x=520, y=30)

        self.medianRadio = Radiobutton(root, font=self.fontRadio,  text="Median", value="median", variable=self.normVar)
        self.medianRadio.place(x=650, y=30)

        self.TMMRadio = Radiobutton(root, font=self.fontRadio, text="TMM", value="TMM", variable=self.normVar)
        self.TMMRadio.place(x=740, y=30)

        ####### Statistics method chosen! ######
        Label(self.frame, font=Font(family="Times New Roman", size=12), text="Statistical Method").place(x = 610, y = 70)

        self.statisticVar = StringVar()
        self.LMMRadio = Radiobutton(root, font=self.fontRadio,  text="Linear mix model", value="LMM", variable=self.statisticVar)
        self.LMMRadio.select()
        self.LMMRadio.place(x=520, y=95)

        self.ttestRadio = Radiobutton(root, font=self.fontRadio,  text="Unpaired t-test", value="ttest", variable=self.statisticVar)
        self.ttestRadio.place(x=680, y=95)

        ####### Browse label and button ######
        self.browseLabel = Label(self.frame, font=Font(family="Times New Roman", size=12), text="Please, choose a PSMs:")
        self.browseLabel.place(x = 80, y = 30)
        self.browseButton = Button(self.frame, text="Browse", justify=LEFT, font=Font(family="Times New Roman", size=12, weight='bold'), command=self.browse)
        self.browseButton.place(x = 240, y = 25)

        Label(self.frame, text="Output Name:", font=Font(family="Times New Roman", size=12)).place(x=80, y=85)
        self.outputNamebox = ScrolledText(self.frame, font=Font(family="Times New Roman", size=12), bd=2)
        self.outputNamebox.place(x=180, y=80, width=320, height=40)

        Label(self.frame, text="Conditions:", font=Font(family="Times New Roman", size=12)).place(x=100, y=155)
        self.conditionbox = ScrolledText(self.frame, font=Font(family="Times New Roman", size=12), bd=2)
        self.conditionbox.place(x=180, y=140, width=500, height=60)

        condtionsFromText = open('condtions.txt').read()
        self.conditionbox.insert(END, condtionsFromText)

        self.browseButtonCondition = Button(self.frame, text="Browse", justify=LEFT, font=Font(family="Times New Roman", size=12, weight='bold'), command=self.browse_condition)
        self.browseButtonCondition.place(x = 700, y = 147)

        Label(self.frame, text="Pairs:", font=Font(family="Times New Roman", size=12)).place(x=130, y=230)
        self.pairsbox = ScrolledText(self.frame, font=Font(family="Times New Roman", size=12), bd=2)
        self.pairsbox.place(x=180, y=220, width=320, height=60)

        pairsFromText = open('pairs.txt').read()
        self.pairsbox.insert(END, pairsFromText)

        self.browseButtonPairs = Button(self.frame, text="Browse", justify=LEFT, font=Font(family="Times New Roman", size=12, weight='bold'), command=self.browse_pairs)
        self.browseButtonPairs.place(x = 520, y = 230)

        self.statusbar = ScrolledText(self.frame, state='disabled')
        self.statusbar.place(x=100, y=300, width=670, height=180)

        self.runbutton = Button(self.frame, text='RUN', fg='black', bg='#b4e67e',
                                font=Font(family="Times New Roman", size=18, weight='bold'), command=self.runbutton_click)
        self.runbutton.place(x=250, y=510, width=150, height=50)

        self.openbutton = Button(self.frame, text='Open', fg='black', bg='#FF5733',
                                font=Font(family="Times New Roman", size=18, weight='bold'), command=self.open_click)
        self.openbutton.place(x=450, y=510, width=150, height=50)
        self.openbutton.configure(state='disabled')

        self.frame.pack()

        self.update_status_box('\n\t\t >> :: mePROD LMM Bot Started! :: <<\n')
        self.update_status_box('\n------------------------------------------------------------------------------\n')

    def Message(self, title, message):
        messagebox.showinfo(title=title, message=message)

    def update_status_box(self, text):
        self.statusbar.configure(state='normal')
        self.statusbar.insert(END, text)
        self.statusbar.see(END)
        self.statusbar.configure(state='disabled')

    def clear_status_box(self):
        self.statusbar.configure(state='normal')
        self.statusbar.delete(1.0, END)
        self.statusbar.see(END)
        self.statusbar.configure(state='disabled')

    def check_main_thread(self):
        root.update()
        if self.myThread.is_alive():
            root.after(1000, self.check_main_thread)
        else:
            self.x = True

    def open_click(self):
        os.startfile(f'{self.outputLocationPath}/{self.outputLocation.strip()}.xlsx')

    def browse(self):
        self.filename = filedialog.askopenfile(parent=self.frame, mode='rb', title='Choose a file')
        self.filenamePretify = str(self.filename).split('/')[-1].split("'>")[0]
        if self.filenamePretify == "None":
            self.Message('Error!', 'Please choose a file!')
            return 0
        self.update_status_box(f'\n"{self.filenamePretify}" file is chosen! \n')

        self.outputLocationPath =  str(self.filename).split("'")[1].replace(str(self.filename).split("'")[1].split("/")[-1],'')

        # self.outputLocationPretify = str(self.outputLocation).split('/')[-1].split("'>")[0]

    def browse_condition(self):
        self.filename_condition = filedialog.askopenfile(parent=self.frame, mode='rb', title='Please, choose a condition text file.')
        self.filenamePretify_condition = str(self.filename_condition).split('/')[-1].split("'>")[0]
        if self.filenamePretify_condition == "None":
            self.Message('Error!', 'Please choose a file!')
            return 0
        self.outputLocationPath_condition =  str(self.filename_condition).split("'")[1].replace(str(self.filename_condition).split("'")[1].split("/")[-1],'')
        condtionsFromText = str(open(self.outputLocationPath_condition+self.filenamePretify_condition).read()).strip()
        self.conditionbox.delete('1.0', END)
        self.conditionbox.insert(END, condtionsFromText)

    def browse_pairs(self):
        self.filename_pairs = filedialog.askopenfile(parent=self.frame, mode='rb', title='Please, choose a pairs text file.')
        self.filenamePretify_pairs = str(self.filename_pairs).split('/')[-1].split("'>")[0]
        if self.filenamePretify_pairs == "None":
            self.Message('Error!', 'Please choose a file!')
            return 0
        self.outputLocationPath_pairs =  str(self.filename_pairs).split("'")[1].replace(str(self.filename_pairs).split("'")[1].split("/")[-1],'')
        pairsFromText = str(open(self.outputLocationPath_pairs+self.filenamePretify_pairs).read()).strip()
        self.pairsbox.delete('1.0', END)
        self.pairsbox.insert(END, pairsFromText)

    def runbutton_click(self):
        #self.update_status_box(f'\n "{self.varDecision.get()}" is selected! \n')
        self.runbutton.configure(state='disabled')
        self.myThread = Thread(target=self.engine)
        self.myThread.daemon = True
        self.myThread.start()
        root.after(1000, self.check_main_thread)

    # def timer(self):
    #     def wrapper(*args, **kwargs):
    #         start_time = time.time()
    #         result = self(*args, **kwargs)
    #         end_time = time.time()
    #         elapsed_time = (end_time - start_time) / 60  # in minutes
    #         print(f"'{self.__name__}' function took {elapsed_time:.2f} minutes to run.")
    #         return result
    #     return wrapper

    def reportAndExport(self, details, data, outputLocation):
        wb = Workbook()
        ws = wb.active
        ws.title = "Info"

        # Given program title
        program_title = f"mePROD LMM App {__version__} by S. Bozkurt @2023"

        # Moving program title to the top
        # ws.merge_cells('H1:I1')
        ws['H1'] = program_title
        ws['H1'].font = sty.Font(size=18, bold=True, color="0072BB")
        ws['H1'].alignment = sty.Alignment(horizontal="center", vertical="center")

        # Splitting the details between columns H and I
        details_split = [(key, value) for key, value in details.items()]

        # Adding the details starting from cell H3
        for index, (label, value) in enumerate(details_split, start=3):
            ws.cell(row=index, column=8, value=label).alignment = sty.Alignment(horizontal="right")
            ws.cell(row=index, column=8).font = sty.Font(size=16)
            ws.cell(row=index, column=9, value=value).alignment = sty.Alignment(horizontal="left")
            ws.cell(row=index, column=9).font = sty.Font(size=16)

        # Current date
        current_date = datetime.now().strftime('%Y-%m-%d')

        # Adding "Date" and "Processed by" details
        ws["H16"] = "Date:"
        ws["H16"].alignment = sty.Alignment(horizontal="right")
        ws["H16"].font = sty.Font(size=16)
        ws["I16"] = current_date
        ws["I16"].font = sty.Font(size=16)

        ws["H17"] = "Processed by:"
        ws["H17"].alignment = sty.Alignment(horizontal="right")
        ws["H17"].font = sty.Font(size=16)
        ws["I17"] = f"User {os.getlogin()}"
        ws["I17"].font = sty.Font(size=16)

        # Add results to a new sheet
        ws_results = wb.create_sheet("Results")

        # Define the font and alignment for the headers
        header_font = sty.Font(bold=True, size=13)
        center_alignment = sty.Alignment(horizontal='center', vertical='center')

        # Write the column headers
        for c_idx, col_name in enumerate(data.columns, 1):
            cell = ws_results.cell(row=1, column=c_idx, value=col_name)
            cell.font = header_font
            cell.alignment = center_alignment

        # Write the data rows
        for r_idx, row in enumerate(data.iterrows(), 2):  # Start from the second row since headers are in the first row
            for c_idx, value in enumerate(row[1], 1):
                ws_results.cell(row=r_idx, column=c_idx, value=value)

        # Specify the path where you'd like to save this workbook
        wb.save(outputLocation)

        return outputLocation

    # @timer # timer decorator for engine function to calculate how long it takes to run!
    def engine(self):
        try:
            if '.xlsx' in self.filenamePretify:
                self.fileRead = pd.read_excel(self.outputLocationPath+self.filenamePretify)
            elif '.txt' in self.filenamePretify:
                self.fileRead = pd.read_csv(self.outputLocationPath+self.filenamePretify,sep='\t',header=0)
            self.update_status_box(f'\n The file is reading..! \n')
        except Exception as e:
            self.update_status_box(f'\n Please choose a file before run! \n')
            self.Message('Error!', 'An Error Occured, please choose a file before run!')
            self.runbutton.configure(state='normal')
            return 0

        self.update_status_box(f'\n The file is read! \n')

        # conditions = ['Light', '0DMSO', '0DMSO', '0DMSO', 'Rotenone', 'Rotenone', 'Rotenone', 'Antimycin', 'Antimycin', 'Antimycin', 'Boost']
        self.conditions = self.conditionbox.get("1.0", END)

        condtionsFile = open(f'{self.outputLocationPath}/condtions.txt','w')
        condtionsFile.write(self.conditions)
        condtionsFile.close()

        condtionsFile = open(f'condtions.txt', 'w') # save the new conditions file to the same directory with the script
        condtionsFile.write(self.conditions)
        condtionsFile.close()

        conditionsFinal = self.conditions.split(',')
        conditionsFinal[-1] = conditionsFinal[-1].strip()

        # pairs = [['0DMSO', 'Rotenone'], ['0DMSO', 'Antimycin']]
        self.pairs = self.pairsbox.get("1.0", END)

        pairsFile = open(f'{self.outputLocationPath}/pairs.txt','w')
        pairsFile.write(self.pairs)
        pairsFile.close()

        pairsFile = open(f'pairs.txt', 'w') # save the new pairs file to the same directory with the script
        pairsFile.write(self.pairs)
        pairsFile.close()

        pairsFinal = self.pairs.split(';')
        pairsFinal[-1] = pairsFinal[-1].strip()
        pairsFinal = [i.strip() for i in pairsFinal]
        pairsFinal = [i.lstrip() for i in pairsFinal]
        pairsFinal = [pairs.split('/') for pairs in pairsFinal]

        normalization_type = self.normVar.get()
        finalStatisticalMethod = self.statisticVar.get().strip()

        if pairsFinal == [['']]:
            finalStatisticalMethod = None
            pairsFinalOutput = 'None' # just for export saying it is none
        else:
            pairsFinalOutput = self.pairs.strip()

        self.update_status_box(f'\n Conditions: {self.conditions.strip()} \n')

        self.update_status_box(f'\n Pairs: {pairsFinalOutput} \n')

        self.update_status_box(f'\n Normalization: {normalization_type.strip()} \n')

        self.update_status_box(f'\n Statistics: {str(finalStatisticalMethod).strip()} \n')

        self.update_status_box(f'\n Running..! \n')

        try:
            randomReportName = f'reports_{str(random.randint(1,100000))}' # random report name that will be used for the report for temporary
            mePROD_class = mePROD(self.outputLocationPath, randomReportName)
            self.data = mePROD_class.engine(self.fileRead, conditionsFinal, pairsFinal, normalization_type, finalStatisticalMethod)
        except Exception as e:
            self.runbutton.configure(state='normal')
            self.update_status_box(f'\n Error is "{e}"! \n')
            self.Message('Error!', f'An Error Occured, please fix it and rerun or contact developer via {__email__ }!')
            return 0

        try:
            if self.data == 0:
                self.update_status_box(f'\n Error is Baseline channel! \n')
                self.Message('Error!', 'Please provide light/baseline channel!')
                self.runbutton.configure(state='normal')
                return 0
        except:
            pass

        self.update_status_box(f'\n Completed..! \n')
        self.update_status_box(f'\n Data is saving..! \n')

        self.outputLocation = self.outputNamebox.get("1.0", END)

        try:
            self.data.to_excel(f'{self.outputLocationPath}/{self.outputLocation.strip()}.xlsx', index=False, engine="openpyxl")

            self.datamePROD = pd.read_excel(f'{self.outputLocationPath}/{self.outputLocation.strip()}.xlsx')
            
            # this function assign and updates gene names
            self.data = mePROD_class.GeneNameEngine(self.datamePROD)
            
            # this function determines mitochondrial proteins from mitocarta 3.0
            self.data = mePROD_class.mito_human(self.data)
            
            # this function assign significantlly changed proteins and add + sign for significant ones p < 0.05
            self.data = mePROD_class.significantAssig(self.data)
            
            # export all data to excel
            # self.data.to_excel(f'{self.outputLocationPath}/{self.outputLocation.strip()}.xlsx', index=False,
            #                    engine="openpyxl")

            # finding back the reposts.txt file and extract the required values
            file_path = os.path.join(self.outputLocationPath, f"{randomReportName}.txt")
            
            # Read the file content
            with open(file_path, "r") as file:
                content = file.read()
            
            # Extract the required values using regular expressions
            totalPeptides = re.search(r"The number of total peptides: (\d+)", content).group(1)
            heavyPeptides = re.search(r"The number of heavy peptides: (\d+)", content).group(1)
            mitosHeavyPeptides = re.search(r"The number of mitochondrial heavy peptides: (\d+)", content).group(1)
            HeavyProteins = re.search(r"The number of heavy proteins: (\d+)", content).group(1)
            mitosHeavyProteins = re.search(r"The number of mitochondrial heavy proteins: (\d+)", content).group(1)

            # Delete the file
            os.remove(file_path)

            details = {
                "Version of the program:": f"{__version__}",
                "The number of total peptides:": totalPeptides,
                "The number of heavy peptides:": heavyPeptides,
                "The number of mitochondrial heavy peptides:": mitosHeavyPeptides,
                "The number of heavy proteins:": HeavyProteins,
                "The number of mitochondrial heavy proteins:": mitosHeavyProteins,
                "":"", # this is for empty line
                "Input file:": self.filenamePretify.strip(),
                "Conditions:": self.conditions.strip(),
                "Pairs:": pairsFinalOutput,
                "Normalization:": normalization_type.strip(),
                "Statistics:": str(finalStatisticalMethod)
            }
            
            self.reportAndExport(details, self.data, f'{self.outputLocationPath}/{self.outputLocation.strip()}.xlsx') # this function creates a report for the data
            
            self.update_status_box(f'\n Saved as {self.outputLocation.strip()}! \n')
            self.Message('Finished!', 'Application Completed!')
            self.openbutton.configure(state='normal')
            self.runbutton.configure(state='normal')
        except Exception as e:
            self.update_status_box(f'\n Error is "{e}"! \n')
            self.Message('Error!', f'An Error Occured, please fix it and rerun or contact developer via {__email__ }!')
            self.runbutton.configure(state='normal')

if __name__ == '__main__':

    root = Tk()
    # root.iconbitmap('files//icon.ico')

    root.title(f"mePROD LMM App {__version__} by S. Bozkurt @2023")
    root.geometry("840x560")
    root.resizable(0, 0)

    # get screen width and height
    ws = root.winfo_screenwidth()  # width of the screen
    hs = root.winfo_screenheight()  # height of the screen

    # calculate x and y coordinates for the Tk root window
    x = (ws / 2) - (840 / 2)
    y = (hs / 2) - (600 / 2)

    # set the dimensions of the screen
    # and where it is placed
    root.geometry('%dx%d+%d+%d' % (840, 600, x, y))

    # root.iconphoto(False, tkinter.PhotoImage(file='icon.ico'))
    MyWindow(root)
    root.wm_iconbitmap('files//icon.ico')
    root.mainloop()